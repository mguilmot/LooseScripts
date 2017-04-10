'''
    Building access list based on AD groups
    Second version, simplified, working with dictionaries, then build the ACL
    Same as the other script, but handles if you have AD groups with different suffixes for RO and RW groups
    ex: 
        ADgroupName-Usr = RO
        ADgroupName-Adm = RW    
    Suffixes have the same length!
    Wrote this for my specific situation. Adapt at will :)
    
    Uses:
    - openpyxl (pip install openpyxl)
    - psutil (pip install psutil)
    - pyad (pip install https://github.com/zakird/pyad/archive/master.zip)
    
    TODO:
    - search for AD groups (instead of input txt file)
    - simplify even further (2 dictionaries instead of 3)
'''

import os, csv, time, psutil
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from pyad import *

### Vars
fileName_in = "c:\\temp\in.txt"                         # Filename containing the AD groups we should check
fileName_csv = "c:\\temp\out.csv"                       # Output CSV filename
fileName_xl = "c:\\temp\\out.xlsx"                      # Output Excel filename
generate_xl = True                                      # Do you want to generate the Excel file
start_time = time.time()                                # Start time
suffixignore="-share"                                   # Suffix of AD groups we should ignore
suffixRO="-usr"                                         # Suffix for AD groups giving read only access
suffixRW="-adm"                                         # Suffix for AD groups giving read + write access
suffixLen=len(suffixRO)                                 # Length of the suffixes
header_row = ['Access List','User ID']                  # Header row1 for our CSV file
header_row2 = ["Group description","AD group prefix"]   # Header row2 for our CSV file
delimiter = ","                                         # Delimiter for our CSV file
dictAllGroups = {}                                      # Placeholder
dictAllUsers = {}                                       # Placeholder
dictAllUsersInfo = {}                                   # Placeholder
    
def printmsg(text="",lines=False):
    '''
        Simple function to print the text we want as output.
    '''
    c = len(text)
    if lines == True:
        l = c*"-"
        return text + "\n" + l + "\n"
    else:
        return "\n" + text + "\n"

def readFile(fileName=fileName_in,suffixignore=suffixignore):
    '''
        Reads the input file, returns them 1 by 1
        Empty lines and lines starting with # are ignored
    '''
    
    with open(fileName) as f:
        for l in f:
            line = l.strip("\n").lower()
            if (line.startswith("#") or len(line)<2) or (suffixignore!="" and line.endswith(suffixignore)) or (suffixignore!="" and line.endswith(suffixignore.lower())):
                continue
            else:
                yield line

def retDispName(user="Administrator"):
    try:
        userObj = pyad.aduser.ADUser.from_cn(user)
        return userObj.get_attribute("displayName")[0]
    except:
        return "None"

def writeCsv(fileName_csv=fileName_csv,delimiter=delimiter,header_row=header_row,header_row2=header_row2,dictAllUsers=dictAllUsers,suffixLen=suffixLen,suffixRO=suffixRO,suffixRW=suffixRW):
    '''
        Writes the desired information to CSV file.
    '''
    
    # dictAllGroupsInfo will be {"filename without suffix without .txt":["description",[RO access list],[RW access list]]}
    dictAllGroupsInfo = {}

    with open(fileName_csv,"w") as f:
        writer = csv.writer(f,delimiter=delimiter,dialect="excel",lineterminator="\n")
        
        print(printmsg(text="Generating CSV file."))
        
        # Print header rows
        writer.writerow(header_row)
        writer.writerow(header_row2)
        
        # Going over dictAllGroups, creating dictAccess to combine RO + RW groups
        for groupName in dictAllGroups:
            groupNameShort = groupName[:-suffixLen]
            groupNameSuffix = groupName[-suffixLen:]
        
            if groupNameShort in dictAllGroupsInfo:
                # Group pair was already checked once, we need to add the other part (RO/RW)
                if groupNameSuffix.lower() == suffixRO.lower():
                    dictAllGroupsInfo[groupNameShort][1] = dictAllGroups[groupName][1]
                else:
                    dictAllGroupsInfo[groupNameShort][2] = dictAllGroups[groupName][1]
                # Finding and removing users that are in both RO and RW groups
                setCommon = set.intersection(set(dictAllGroupsInfo[groupNameShort][1]),set(dictAllGroupsInfo[groupNameShort][2]))
                for doubleUser in setCommon:
                    dictAllGroupsInfo[groupNameShort][1].remove(doubleUser)
            else:
                dictAllGroupsInfo[groupNameShort] = [dictAllGroups[groupName][0],[],[]]
                if groupNameSuffix.lower() == suffixRO.lower():
                    dictAllGroupsInfo[groupNameShort][1] = dictAllGroups[groupName][1]
                else:
                    dictAllGroupsInfo[groupNameShort][2] = dictAllGroups[groupName][1]
                    
        # Writing the actual lines in the CSV
        for item in sorted(list(dictAllGroupsInfo.keys())):
        
            linetxt = [""] * 2
            usersinfile = [""] * len(dictAllUsers) 
            
            linetxt[0] = dictAllGroupsInfo[item][0][0]  # group desc
            linetxt[1] = item                           # group name without suffix
            
            for userRO in dictAllGroupsInfo[item][1]:
                userlocation = dictAllUsers[userRO.upper()]
                usersinfile[userlocation]="RO"
            for userRW in dictAllGroupsInfo[item][2]:
                userlocation = dictAllUsers[userRW.upper()]
                usersinfile[userlocation]="RW"
            
            writer.writerow(linetxt+usersinfile)
        
        writer.writerow([])                             # empty line
        writer.writerow([])                             # empty line
        writer.writerow(["RO = Read Only access"])      # extra info
        writer.writerow(["RW = Read + Write access"])   # extra info
        
        print(printmsg(text="Done. Check " + fileName_csv + " for output."))

def findUsers(groupName="Administrators",dictAllGroups=dictAllGroups,dictAllUsers=dictAllUsers,dictAllUsersInfo=dictAllUsersInfo):
    '''
        Finds users that are members of the AD group, adds them to dictAllUsers if not already in there
        Adds the group and group description and all members to dictAllGroups
    '''
    lstUsers = []
    try:
        grpObj = pyad.adgroup.ADGroup.from_cn(groupName)
        grpMembers = pyad.adgroup.ADGroup.get_members(grpObj)
        users = [str(member).split(",")[0][12:] for member in grpMembers]
        if len(users)>0:
            desc = grpObj.get_attribute("description")[0]
            if desc == "":
                desc="No group description"
            for user in users:
                if user.startswith("="):
                    # user = AD group
                    continue
                else:
                    # Group is not empty, user is not an AD group
                    lstUsers.append(user.upper())
                    if user not in dictAllUsers:
                        displayName = retDispName(user=user)
                        dictAllUsersInfo[user] = displayName
            dictAllGroups[groupName] = [[desc],lstUsers]
            return len(lstUsers)
        else:
            # Group has no members
            return "0. Skipping."
    except:
        # Group does not exist
        return "0. Skipping."

def writeXl(fileName_csv=fileName_csv,fileName_xl=fileName_xl,delimiter=delimiter):
    '''
        Converts our CSV file to XLSX format
    '''
    
    numRows = 0
    
    print(printmsg(text="Converting CSV to XLSX"))
    
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(fileName_csv) as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            ws.append(row)

        # Setting max width
        dims = {}
        for row in ws.rows:
            numRows += 1
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value
        
        # Sorting our columns
        lstCol = list(dims.keys())
        lstCol.sort(key=lambda item: (len(item),item))
        
        # Setting formatting
        print(printmsg(text="Setting formatting in XLSX"))
        for colNum,col in enumerate(lstCol):
            # Running over columns
            if colNum > 1:
                # column width = 3
                ws.column_dimensions[col].width = 4
            for rowNum in range(numRows):
                # Running over rows
                curCell = ws[str(col.upper()) + str(rowNum+1)]
                curCell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                if colNum > 1:
                    # We do not touch the first two columns, borders are already set
                    if rowNum < 2:
                        # Formatting should be: vertical text + all border + center
                        curCell.alignment = Alignment(horizontal='center',text_rotation=90)
                    else:
                        # Formatting should be : all borders + center
                        curCell.alignment = Alignment(horizontal='center')
        # Freeze panes
        ws.freeze_panes = ws["B3"]
        
    wb.save(fileName_xl)

    print(printmsg(text="Done. Check " + fileName_xl + " for output."))
    
# Reading input file: filling dictionaries with requested data
for line in readFile():
    print("Looking for users for:",line,end="")
    print(". Found:",findUsers(groupName=line))

# Filling header rows with our users in alphabetical order. Filling dict to remember their location for our CSV
for num,user in enumerate(sorted(dictAllUsersInfo)):
        dictAllUsers[user]=num
        header_row.append(user)
        header_row2.append(dictAllUsersInfo[user])

# Writing our output files
if len(dictAllUsers)>0:
    writeCsv()
    if generate_xl == True:
        writeXl()
else:
    print("No data. Skipping CSV and XLS.\n")

# Print execution time
print("Execution time: --- %s seconds ---" % '{:02.2f}'.format((time.time() - start_time)))
process = psutil.Process(os.getpid())
print("Memory Usage: --- %s MB ---" % '{:02.2f}'.format((process.memory_info().rss / float(2 ** 20))))
print()